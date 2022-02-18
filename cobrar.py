import win32com.client as win32
from win32com.client.makepy import main
from tkinter import *
from tkinter.filedialog import askopenfilename, askopenfilenames
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import time
import datetime

outlook = win32.Dispatch('outlook.application')
outlook = outlook.GetNamespace('MAPI')
inbox = outlook.GetDefaultFolder(6)
messages = inbox.Items
messages.Sort("[ReceivedTime]", True)

validator_column = 'C'
subject_column = 'D'

Tk().withdraw()
wb_name = askopenfilename()

wb = load_workbook(wb_name, keep_vba=True)
sheet = wb['Cobrar Cotações']


def get_saudacao():
    hora = datetime.datetime.now().ctime()
    hora = int(hora[11:13])
    if hora in range(0, 13):
        saudacao = 'bom dia'
    if hora in range(13, 18):
        saudacao = 'boa tarde'
    if hora in range(18, 24):
        saudacao = 'boa noite'
    return saudacao


for row in range(3, sheet.max_row + 1):
    subject_cell = sheet[f'{subject_column}{row}']
    validator_cell = sheet[f'{validator_column}{row}']
    if subject_cell.value != None and validator_cell.value == 'Não':
        for message in messages:
            if subject_cell.value in message.Subject:
                reply = message.ReplyAll()
                reply.Display()

                TEXT_COLOR = '#3c4064'
                text = f'''
                    <p style='color: {TEXT_COLOR}'p>{get_saudacao().capitalize()}!</p>
                    <p style='color: {TEXT_COLOR}'p>Gentiliza informar previsão do envio do orçamento solicitado.</p>
                    <p style='color: {TEXT_COLOR}'p>Salientamos que precisamos da cotação o quanto antes.</p>
                    <p style='color: {TEXT_COLOR}'p>Agradecemos a atenção e ficamos à disposição.</p>
                '''

                index = reply.HTMLbody.find('>', reply.HTMLbody.find('<body'))
                reply.HTMLbody = reply.HTMLbody[:index +
                                                1] + text + reply.HTMLbody[index + 1:]
                reply.CC = 'comercialenergia@visionsistemas.com.br'
                break
