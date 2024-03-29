from tkinter import *
from tkinter.filedialog import askopenfilename, askopenfilenames
import re
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import datetime
import time
from copy import copy

import win32com.client as win32
from win32com.client.makepy import main


def get_database(wb):
    fornecedores = {}
    for ws_name in wb.sheetnames:
        if 'fornecedor' in ws_name.lower():
            current_working_sheet = wb[ws_name]
            company_name = current_working_sheet['A1'].value.upper()
            fornecedores[company_name] = {}
            voltage_columns = {
                'B': 'UAT',
                'C': 'AT',
                'D': 'MT',
                'E': 'BT'
            }
            for column in voltage_columns:
                d = fornecedores[company_name][voltage_columns[column]] = {}
                for row_num in range(4, 18):
                    equipamento = current_working_sheet[f'A{row_num}'].value
                    fornecimento = current_working_sheet[f'{column}{row_num}'].value
                    d[equipamento] = True if fornecimento == 'Sim' else False

            contatos = fornecedores[company_name]['contatos'] = []
            for row_num in range(4, 18):
                nome = current_working_sheet[f'G{row_num}'].value
                sobrenome = current_working_sheet[f'H{row_num}'].value
                tratamento = current_working_sheet[f'I{row_num}'].value
                email = current_working_sheet[f'J{row_num}'].value

                if nome != None:
                    info = {
                        'nome': nome,
                        'sobrenome': sobrenome,
                        'email': email,
                        'tratamento': tratamento,
                    }
                    contatos.append(info)

    return fornecedores


class Equipamento:
    def __init__(self, tensao, type, qtd, descricao, et=None):
        self.et = et
        self.tensao = tensao
        self.type = type
        self.qtd = qtd
        self.descricao = descricao


class Emailer:
    def __init__(self, receiver, subject, body, attachments=None):
        self.receiver = receiver
        self.subject = subject
        self.body = body
        self.attachments = attachments

    def prepare_emails(self):
        outlook = win32.Dispatch('outlook.application')
        email = outlook.CreateItem(0)
        email.Display()
        bodystart = re.search("<body.*?>", email.HTMLBody)
        email.HTMLBody = re.sub(bodystart.group(), bodystart.group() +
                                self.body, email.HTMLBody)
        email.To = self.receiver
        email.Subject = self.subject
        if self.attachments and len(self.attachments) > 0:
            if None in self.attachments:
                self.attachments.remove(None)
            for attachments in self.attachments:
                for attachment in attachments:
                    email.Attachments.Add(attachment)
        email.CC = 'comercialenergia@visionsistemas.com.br'

    def create_summary_sheet(self, fornecedor, wb):
        sheet = wb['Cobrar Cotações']

        for row in range(3, 100):
            fonecedor_cell = sheet[f'B{row}']
            check_cell = sheet[f'C{row}']
            subject_cell = sheet[f'D{row}']
            if fonecedor_cell.value == None:
                fonecedor_cell.value = fornecedor
                check_cell.value = 'Não'
                subject_cell.value = self.subject
                break


def get_et():
    window = Tk()
    window.withdraw()
    et = askopenfilenames()
    window.destroy()
    if et != '':
        return et
    return None


def get_data_from_equipamentos_sheet(wb):
    window = Tk()
    window.eval('tk::PlaceWindow . center')
    window.withdraw()
    window.destroy()
    equipamentos = []

    current_working_sheet = wb['Equipamentos']
    already_have_attachments = []

    descricao_column = 'B'
    voltage_column = 'D'
    qtd_column = 'C'
    eq_type_column = 'E'

    for row in range(3, current_working_sheet.max_row):
        qtd_cell_value = current_working_sheet[f'{qtd_column}{row}'].value
        if qtd_cell_value != None:
            descricao_cell_value = current_working_sheet[f'{descricao_column}{row}'].value
            voltage_cell_value = current_working_sheet[f'{voltage_column}{row}'].value
            eq_type_cell_value = current_working_sheet[f'{eq_type_column}{row}'].value

            if eq_type_cell_value not in already_have_attachments:
                root = Tk()
                root.geometry('300x50+0+0')
                Label(
                    root, text=f'Insira os anexos para: {eq_type_cell_value}').pack()
                et = get_et()
                root.destroy()
                already_have_attachments.append(eq_type_cell_value)

            new_eq = Equipamento(voltage_cell_value,
                                 eq_type_cell_value, qtd_cell_value, descricao_cell_value, et)

            equipamentos.append(new_eq)

    fornecedores = get_database(wb)
    return equipamentos, fornecedores


class Main_widget():

    def __init__(self):
        self.root = Tk()
        self.root.eval('tk::PlaceWindow . center')
        self.num_proposta = ''
        self.dias = ''
        self.root.title('Disparador de Cotações')
        Label(self.root, text='Número da Proposta:').pack()
        self.e1 = Entry(self.root)
        self.e1.pack()
        Label(self.root, text='Quatidade de dias para receber a cotação:').pack()
        self.e2 = Entry(self.root)
        self.e2.pack()
        b = Button(text='Pronto', command=self.handle_click)
        b.pack()
        self.root.mainloop()

    def handle_click(self):
        self.num_proposta = self.e1.get()
        self.dias = self.e2.get()
        self.root.destroy()


def make_fornecedores_resumo(wb):
    equipamentos, FORNECEDORES = get_data_from_equipamentos_sheet(wb)
    resumo = {}
    for fornecedor in FORNECEDORES:
        for equipamento in equipamentos:
            try:
                validator = FORNECEDORES[fornecedor][equipamento.tensao][equipamento.type]
            except Exception as e:
                print(e)
            if validator:
                if fornecedor not in resumo:
                    resumo.update(
                        {fornecedor: []})
                resumo[fornecedor].append(equipamento)

    return resumo, FORNECEDORES


def write_fornecedor(eq, fornecedor, wb):
    title_row = 3
    current_working_sheet = wb[eq.type.capitalize()]
    current_working_sheet.sheet_state = 'visible'
    style = copy(
        current_working_sheet['B3']._style)

    for column in range(5, current_working_sheet.max_column + 1):
        current_cell = current_working_sheet[f'{get_column_letter(column)}{title_row}']
        if current_cell.value == None:
            current_cell.value = fornecedor.upper()
            current_cell._style = style
            break


def build():
    main_widget = Main_widget()

    root = Tk()
    root.geometry('300x50+0+0')
    Label(root, text=f'Selecione a planilha "Equipamentos"').pack()
    wb_name = askopenfilename()
    Tk().withdraw()
    root.destroy()
    wb = load_workbook(wb_name, keep_vba=True)

    '''Limpa a sheet de cobrar cotações'''
    sheet = wb['Cobrar Cotações']
    for row in range(3, 100):

        fonecedor_cell = sheet[f'A{row}']
        check_cell = sheet[f'B{row}']
        subject_cell = sheet[f'C{row}']
        fonecedor_cell.value = None
        check_cell.value = None
        subject_cell.value = None

    resumo, FORNECEDORES = make_fornecedores_resumo(wb)

    for fornecedor in resumo:
        contatos = FORNECEDORES[fornecedor]['contatos']
        hora = datetime.datetime.now().ctime()
        hora = int(hora[11:13])

        if hora in range(0, 13):
            introducao = 'bom dia'
        if hora in range(13, 18):
            introducao = 'boa tarde'
        if hora in range(18, 24):
            introducao = 'boa noite'

        if len(contatos) > 1:
            tratamento = 'Prezados'
        if len(contatos) == 1:
            tratamento = contatos[0]['tratamento'].capitalize() + \
                ' ' + contatos[0]['nome']

        date = str(datetime.date.today() +
                   datetime.timedelta(days=int(main_widget.dias)))

        year = date[0:4]
        month = date[5:7]
        day = date[8:]
        date = f'{day}/{month}/{year}'


        text = ''
        item = 1
        ets = []
        equipamentos = []

        for eq in resumo[fornecedor]:
            write_fornecedor(eq, fornecedor, wb)
            if eq.type.capitalize() not in equipamentos:
                equipamentos.append(eq.type.capitalize())

            if eq.et not in ets:
                ets.append(eq.et)

            TEXT_COLOR = '#3c4064'
            TABLE_STYLE = 'border-collapse:collapse; margin: 5px; font-size: 14px; width: 500px;'
            TR_STYLE = 'background-color:#154c79; color: white; font-weight:bold;border: 1px solid; border: 1px solid'
            TH_STYLE = 'padding: 1px 4px; border: 1px solid'
            TD_STYLE = 'text-align:center; padding: 1px 4px; border: 1px solid; background-color: white'

            text += f"""
            <tr'>
                <td style='{TD_STYLE}'>{item}</td>
                <td style='{TD_STYLE}'>{eq.descricao}</td>
                <td style='{TD_STYLE}'>{eq.qtd}</td>
            </tr>
            """
            item += 1
            body = f"""
                    <div>
                    <p style='color: {TEXT_COLOR}'p>{tratamento}, {introducao}, </p>
                    <p style='color: {TEXT_COLOR}'p>Devido à necessidade do setor comercial da Vision Engenharia, solicitamos o orçamento dos equipamentos destacados na tabela a seguir. O anexo contém mais informações a serem consideradas.</p>
                    <table style='{TABLE_STYLE}'>
                            <tr style='{TR_STYLE}'>
                                <th style='{TH_STYLE}'>ITEM</th>
                                <th style='{TH_STYLE}'>DESCRIÇÃO</th>
                                <th style='{TH_STYLE}'>QTD</th>
                            </tr>
                        {text}   
                    </table>
                        <p style='color: {TEXT_COLOR}'>O prazo máximo para a cotação é dia {date}.</p>
                        <p style='color: {TEXT_COLOR}'>Observações:</p>
                        <p style='color: {TEXT_COLOR}'>- Caso não seja possível realizar o orçamento dentro do prazo, favor informar em resposta a este e-mail;</p>
                        <p style='color: {TEXT_COLOR}'>- Favor responder a todos os que estão em cópia.</p>
                        <p style='color: {TEXT_COLOR}'>Desde já, a Vision agradece o seu apoio e a sua atenção e nos colocamos à disposição para sanar quaisquer dúvidas sobre o orçamento.</p>
                    </div>
                """

        to = []
        for contato in contatos:
            to.append(contato['email'])
        to = ';'.join(to)

        equipamentos = ', '.join(equipamentos)

        subject = f'0006 | RFQ{main_widget.num_proposta} - Cotação {equipamentos} [{fornecedor}]'
        emailer = Emailer(to, subject, body, ets)
        emailer.prepare_emails()

        emailer.create_summary_sheet(fornecedor, wb)

        wb.save(wb_name)


build()
